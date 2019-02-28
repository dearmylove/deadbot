import threading
import discord
import asyncio
import random
import openpyxl
import datetime
import Image
import UnSae


client = discord.Client()


def Running(): # 10분 간격으로 'Bot_is_Running' 이라는 문자열을 출력하며 봇의 상태를 알려줌.
    print('Bot_is_Running')
    threading.Timer(600, Running).start()

Running()


@client.event
async def on_member_join(member): # 방에 막 들어온 Member에게 "새로 오신 분!" 이라는 Role을 부여함
    role = ""
    for i in member.server.roles:
        if i.name == "새로 오신 분!":
            role = i
            break
    await client.add_roles(member, role)


@client.event
async def on_ready(): # 봇의 로그인을 알림
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("======================")
    await client.change_presence(game=discord.Game(name='!명령어로 봇의 명령어 확인', type=1))


@client.event
async def on_message(message):
    ###### 봇의 Preset ######
    ###### 봇의 Preset ######


    if message.content.startswith("!명령어"):
        embed = discord.Embed(
            title='꿀바데 봇의 명령어 입니다.',
            description='참고하시고 클린한 서버 만들어 나갑시다!',
            colour=discord.Colour.blue()
        )
        embed.add_field(name='!응삼이', value=': 트위치 스트리머 응삼이에 대해 알려줍니다.', inline=False)
        embed.add_field(name='!출첵', value=': 매일매일 24시간에 한 번 출석체크가 가능합니다.', inline=False)
        embed.add_field(name='!출확', value=': 몇 번 출석했는지 확인 할 수 있습니다.', inline=False)
        embed.add_field(name='!선택 항목', value=": 여러 항목 중 하나를 뽑을 수 있습니다.", inline=False)
        embed.add_field(name='!정보', value=': 본인의 디스코드 정보에 대해 알려줍니다..', inline=False)
        embed.add_field(name='!프로필 스팀이름', value=": 스팀에서 프로필을 검색할 수 있습니다.", inline=False)
        embed.add_field(name='!사진 입력내용(띄어쓰기 X)', value=": 네이버의 사장 상위 사진을 보여줍니다.", inline=False)
        embed.add_field(name='!건의 건의내용', value =': 추가됐으면 좋겠다는 기능을 건의할 수 있습니다.', inline = False)
        embed.add_field(name='!운세 띠이름', value=': 오늘의 운세를 확인할 수 있습니다. \n띠 목록 : 소띠, 쥐띠, 호랑이띠, 토끼띠, 용띠, 뱀띠, 말띠, 양띠, 원숭이띠, 닭띠, 개띠, 돼지띠', inline=False)


        await client.send_message(message.channel, embed=embed)

#    if message.content.startswith("!카즈"):
#        if message.content.startswith("!카즈바보"):
#            await client.send_message(message.channel, "카바즈보")
#        else:
#            await client.send_message(message.channel, "카즈 변태래요~")

    if message.content.startswith("!응삼이"):
        embed = discord.Embed(
            title='데바데 3000시간 고인물 응삼이!!',
            description='데바데 등 종합게임스트리머, 매주 수요일 휴방. 오후 5시 방송 시작',
            colour=discord.Colour.orange()
        )
        embed.set_author(name='응삼이', icon_url='https://static-cdn.jtvnw.net/jtv_user_pictures/0bad62f6-882d-4ecd-a8c9-1b7e5a1ba627-profile_image-70x70.png')
        embed.add_field(name='트위치 생방송을 보러가고 싶다면?', value=': https://www.twitch.tv/dmdtkadl69', inline=False)
        embed.add_field(name='실력/꿀잼 편집 영상을 다시보고 싶다면?', value=': https://www.youtube.com/channel/UCifzX6G1ONZtV45qxHdE1Xg', inline=False)
        embed.add_field(name='응삼이와 소통하고 싶다면?', value=': https://tgd.kr/dmdtkadl69', inline=False)
        embed.add_field(name='응삼이에게 후원하고 싶다면?', value=": https://twip.kr/dmdtkadl69", inline=False)

        await client.send_message(message.channel, embed=embed)

    if message.content.startswith("!공간이동"):
        channel = discord.utils.get(client.get_all_channels(), id='529712128177995787') # 출첵 - 이모지신청
        print(channel)
        await client.send_message(channel, "TEST")

    if message.content.startswith("!사진"):
        img = message.content.split(" ")
        realimg = img[1:]
        name = ""
        for i in realimg:
            name = name + i + " "
        imgsrc = Image.get_image(name)
        print(imgsrc)
        await client.send_message(message.channel, imgsrc)

    if message.content.startswith("!정보"):
        id = message.author.id
        date = datetime.datetime.utcfromtimestamp(((int(id) >> 22) + 1420070400000) / 1000)

        embed = discord.Embed(colour=discord.Colour.green())
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일",
                        inline=True)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)

        await client.send_message(message.channel, embed=embed)

    if message.content.startswith("!운세"):
        text = message.content.split(" ")
        textsrc = UnSae.get_image(text[1])
        print(textsrc)
        await client.send_message(message.channel, textsrc)

    if message.content.startswith("!토론신청"):

        profile = message.author.display_name
        print(profile)
        client.send_message(client.start_private_message(message.author), profile + "참가신청")

    if message.content.startswith("!건의"):
        content = message.content.split(" ")
        content = content[1:]
        name = ""
        for i in content:
            name = name + i + " "
        print(name)
        dalcom = discord.utils.get(client.get_all_members(), id='257098084352393217')
        print(dalcom)
        # 413714272120602624)
        await client.send_message(dalcom, message.author.name + "(" + message.author.display_name + ")" + " : " + name)

    if message.content.startswith("!출첵"):
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")
        sheet = file.active

        for i in range(1, 298):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                if int(sheet["B" + str(i)].value) <= int(datetime.datetime.now().strftime("%Y%m%d%H%M%S")):
                    await client.send_message(message.channel, "출석체크 완료!")
                    a = datetime.datetime.now() + datetime.timedelta(hours = 24)
                    sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                    sheet["C" + str(i)].value += 1

                    file.save("쿨타임(쳌).xlsx")
                    break
                else:
                    checktime = sheet["B" + str(i)].value
                    checktime = str(checktime)
                    year = int(checktime[0:4])
                    month = int(checktime[4:6])
                    day = int(checktime[6:8])
                    hour = int(checktime[8:10])
                    minute = int(checktime[10:12])
                    second = int(checktime[12:14])
                    plan_time = datetime.datetime(year, month, day, hour, minute, second)
                    print(plan_time)
                    tdy = datetime.datetime.now()
                    time_remainer = plan_time - tdy
                    time_remainer = time_remainer.seconds
                    time_remainer_hour = int(time_remainer / 3600)
                    time_remainer_minute = int((time_remainer - (time_remainer_hour * 3600)) / 60)
                    time_remainer_second = int(time_remainer - (time_remainer_hour * 3600) - (time_remainer_minute * 60))

                    await client.send_message(message.channel, "남은 시간 : {}시간 {}분 {}초. 예상날짜 : {}일 {}시 {}분 {}초".format(time_remainer_hour, time_remainer_minute, time_remainer_second, day, hour, minute, second))
                    break

            if str(sheet["A" + str(i)].value) == '-':
                sheet["A" + str(i)].value = str(message.author.id)
                a = datetime.datetime.now() + datetime.timedelta(hours = 24)
                sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                sheet["C" + str(i)].value += 1

                file.save("쿨타임(쳌).xlsx")
                await client.send_message(message.channel, "출석체크 완료!")
                break

    if message.content.startswith("!출확"):
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")
        many = message.content.split(" ")
        name = many[1:]  # 사람 이름
        length = len(message.author.roles)
        role = message.author.roles[length - 1].name
        sheet = file.active

        for i in range(1, 298):
            if str(sheet["A" + str(i)].value) == message.author.id:
                times = str(sheet["C" + str(i)].value)
                var_time = int(times)
                length = len(message.author.roles)

                if sheet["C" + str(i)].value == 1:
                    member = message.author.id
                    name = ":blue_heart: 청정수"
                    await client.send_message(message.channel, "!등급 " + member + " " + name)







                embed = discord.Embed(
                    title= 'Profile'
                           ,
                    colour=discord.Colour.orange()
                )
                embed.add_field(name=' 플레이어 정보', value= message.author, inline=False)
                embed.add_field(name=' 플레이어 등급', value= role, inline=False)
                embed.add_field(name=' 출석횟수', value='지금까지 ' + times + '번 출석하셨습니다.', inline=False)

                await client.send_message(message.channel, embed=embed)

                file.save("쿨타임(쳌).xlsx")
                break

    if message.content.startswith("!선택"):
        person = message.content.split(" ")
        person = person[1:]
        random.shuffle(person)

        await client.send_message(message.channel, person[1] + "→" + '선택되었습니다!')

    if message.content.startswith("!프로필"):
        content = message.content.split(" ")
        id = content[1]

        await client.send_message(message.channel, "https://steamcommunity.com/search/users/#text=" + id)

    if message.content.startswith("!등급"):  # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림
        role = ":blue_heart: 청정수"
        rolename = message.content.split(" ")
        print("TEST")
        member = discord.utils.get(client.get_all_members(), id=str(rolename[1]))
        print(member)
        roletrue = rolename[2:]
        name = ""
        for i in roletrue:
            name = name + i + " "
        print(name)
        for i in message.server.roles:
            print(i)
            if i.name == name:
                role = i

                break

        print(role)

    if message.content.startswith("!test"):
        msg = message.content.split(" ")
        idea = msg[1]
        member = discord.utils.get(client.get_all_members(), id=idea)
        await client.send_message(message.channel, member)


        # nickname = discord.Member.nick(message.author.id)
        # await client.change_nickname(member, ":blue_heart: " + nickname)
        await client.add_roles(member, role)


"""
    if message.content.startswith("!등급"): # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id = str(rolename[1]))
        roletrue = rolename[2:]
        name = ""
        for i in roletrue:
            name = name + i + " "
        name = name[:-1]
        for i in message.server.roles:
            if i.name == name:
                role = i
                break

        await client.add_roles(member, role)

    if message.content.startswith("!등삭"): # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id=str(rolename[1]))

        roletrue = rolename[2:]
        name = ""
        for i in roletrue:
            name = name + i + " "
        # name = name[:-1]

        for i in message.server.roles:
            if i.name == name:
                role = i
                break

        await client.remove_roles(member, role)
        """

"""
    if message.content.startswith("!출석보정"):
        mes = message.content.split(" ")
        id = mes[1]
        count = int(mes[2])
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")
        sheet = file.active
        for i in range(1, 101):
            if str(sheet["A" + str(i)].value) == str(id):

                sheet["C" + str(i)].value += count
                file.save("쿨타임(쳌).xlsx")
                break
            if str(sheet["A" + str(i)].value) == '-':
                sheet["A" + str(i)].value = str(id)
                sheet["C" + str(i)].value += count

                file.save("쿨타임(쳌).xlsx")

                await client.send_message(message.channel, "출석체크 완료!")
                break
                """






client.run('NTMyNjEzMzU5ODY3MDAyODgw.Dxh69g._AJOXBIwcD9AotBKwjDi-bEAFeE')